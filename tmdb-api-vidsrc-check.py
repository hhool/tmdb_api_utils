import openpyxl
import aiohttp
import asyncio
import argparse

def read_excel_file(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    tmdb_data = []
    
    # Record the original row number when reading
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        tmdb_id = row[1]
        vote = row[5]
        
        # Type conversion
        try:
            vote_int = int(vote) if vote is not None else 0
        except ValueError:
            vote_int = 0
            
        try:
            tmdb_id_int = int(tmdb_id) if tmdb_id is not None else 0
        except ValueError:
            tmdb_id_int = 0
            
        tmdb_data.append((tmdb_id_int, vote_int, row_idx))  # Store row number
    
    # Sort by vote count
    sorted_data = sorted(tmdb_data, key=lambda x: x[1], reverse=True)
    
    # Extract (tmdb_id, original row number)
    processed = [(item[0], item[2]) for item in sorted_data]  # (tmdb_id, original row number)
    
    return wb, ws, processed

async def verify_vidsrc_link(session, tmdb_id, media_type):
    print(f"⏳Checking {media_type} ID: {tmdb_id}")
    try:
        url = f"https://vidsrc.me/embed/{media_type}/{tmdb_id}/" + \
              ("1/1" if media_type == "tv" else "")
              
        async with session.get(url, timeout=10) as response:
            final_url = str(response.url)
            status = response.status
            
            # Validity check
            if status == 200 and "embed" in final_url:
                print(f"✅ Valid: {tmdb_id}")
                return final_url
            print(f"❌ Invalid: {tmdb_id} (Status: {status})")
            return None
    except Exception as e:
        print(f"🚨 Error: {tmdb_id} - {str(e)}")
        return None

async def main(filepath, num_records=None):
    # Determine media type
    media_type = "tv" if "tv_" in filepath else "movie"
    
    # Read file and get sorted data
    wb, ws, processed_data = read_excel_file(filepath)  # processed_data format: [(tmdb_id, original row number)]
    
    processed_data = processed_data[:num_records] if num_records else processed_data

    # Create result dictionary {row number: result URL}
    results = {}
    
    # Start session, and verify each tmdb_id
    # Dump progress to console
    total = len(processed_data)
    print(f"🔍 Verifying {total} records")
    async with aiohttp.ClientSession() as session:
        for idx, (tmdb_id, row_idx) in enumerate(processed_data, 1):
            # Perform verification
            print(f"\n🔍 Processing {idx}/{total}...")
            url = await verify_vidsrc_link(session, tmdb_id, media_type)
            results[row_idx] = url  # Store result by original row number
            
            # Request interval
            # check idx to avoid the last sleep
            # check url to avoid unnecessary sleep
            if idx < len(processed_data):
                if url:
                    print(f"⏳ Waiting 3 seconds...")
                    await asyncio.sleep(3)
                else:
                    print(f"⏳ Failed request. Waiting 1 second...")
                    await asyncio.sleep(1)

    # Locate result column
    header = [cell.value for cell in ws[1]]
    result_col = len(header) + 1  # New column position

    # Add header (if not exists)
    if "MediaUrl" not in header:
        ws.cell(row=1, column=result_col, value="MediaUrl")

    # Write results to corresponding rows
    for row_idx, url in results.items():
        ws.cell(row=row_idx, column=result_col, value=url or "Not Found")

    # Save file
    wb.save(filepath)
    print(f"\n🎉 Results saved to {filepath}")

    # Close session
    await session.close()

    # Dump results to file path, for further analysis
    # file name is similar to the input file name, with a suffix
    dump_file = f"{filepath.split('.')[0]}_dump.txt"
    success_count = sum(1 for url in results.values() if url)
    failed_count = len(results) - success_count
    with open(dump_file, "w") as f:
        f.write(f"Total: {len(results)}\n")
        f.write(f"Success: {success_count}\n")
        f.write(f"Failed: {failed_count}\n")
        f.write(f"Success Rate: {success_count / len(results) * 100:.2f}%\n")
        f.write("\n".join(f"{row_idx}: {url}" for row_idx, url in results.items()))
    print(f"\n📝 Results dumped to {dump_file}")
    # Dump total count and success rate
    print(f"\nTotal: {len(results)}")
    print(f"Success: {success_count}")
    print(f"Failed: {failed_count}")
    print(f"Success Rate: {success_count / len(results) * 100:.2f}%")

# Usage:
# tmdb-api-vidsrc-check.py -f movie_x.xlsx -n 100
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Verify TMDB IDs on vidsrc.me")
    parser.add_argument("-f", "--filepath", required=True, help="Path to the Excel file")
    parser.add_argument("-n", "--num_records", type=int, default=None, help="Number of records to process (default: all)")
    args = parser.parse_args()
    
    asyncio.run(main(args.filepath, args.num_records))